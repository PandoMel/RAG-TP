from pathlib import Path

from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader


def parse_txt(path: Path) -> tuple[str, float]:
    content = path.read_text(encoding="utf-8", errors="ignore")
    score = min(1.0, len(content.strip()) / 5000)
    return content, score


def parse_docx(path: Path) -> tuple[str, float]:
    doc = DocxDocument(path)
    content = "\n".join(par.text for par in doc.paragraphs if par.text)
    score = min(1.0, len(content.strip()) / 5000)
    return content, score


def parse_xlsx(path: Path) -> tuple[str, float]:
    wb = load_workbook(path, read_only=True, data_only=True)
    blocks = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(" | ".join([str(v) if v is not None else "" for v in row]))
        blocks.append(f"## sheet: {ws.title}\n" + "\n".join(rows))
    content = "\n\n".join(blocks)
    score = min(1.0, len(content.strip()) / 5000)
    return content, score


def parse_pdf_builtin(path: Path) -> tuple[str, float]:
    reader = PdfReader(path)
    pages = [page.extract_text() or "" for page in reader.pages]
    content = "\n".join(pages)
    non_empty = sum(1 for p in pages if p.strip())
    score = non_empty / max(1, len(pages))
    return content, score
